#!/usr/bin/env python3
"""
임대시장/B.상업용 공식 xlsx + 상권구획도 → rent_stats.

하위시장 시트만. 104·106(규모별 서울광역) 제외. 109–112는 floor_band로 저장.
"""

from __future__ import annotations

import argparse
import json
import logging
import math
import sys
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from psycopg2.extras import execute_values, Json
from sqlalchemy import text

REPO = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO / "backend"))
sys.path.insert(0, str(REPO / "pipeline"))

from app.rent.sangkwon_agg import (  # noqa: E402
    SHEET_ASSET,
    SHEET_FLOOR_BAND,
    SKIP_SHEETS,
    is_aggregate_name,
    metric_from_item,
    parse_number,
    parse_quarter_header,
)
from rent.db_utils import get_rent_engine  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

DEFAULT_XLSX_DIR = REPO / "임대시장" / "B.상업용"
DEFAULT_SHP = DEFAULT_XLSX_DIR / "상권구획도(업로드용)" / "상권구획도2024.shp"
MERCATOR_R = 6378137.0


def mercator_to_lonlat(x: float, y: float) -> tuple[float, float]:
    lon = (x / MERCATOR_R) * (180.0 / math.pi)
    lat = (2.0 * math.atan(math.exp(y / MERCATOR_R)) - math.pi / 2.0) * (180.0 / math.pi)
    return lon, lat


def _ring_to_lonlat(points: list) -> list[list[float]]:
    out: list[list[float]] = []
    for pt in points:
        if pt is None or len(pt) < 2:
            continue
        lon, lat = mercator_to_lonlat(float(pt[0]), float(pt[1]))
        if out and out[-1][0] == lon and out[-1][1] == lat:
            continue
        out.append([lon, lat])
    if len(out) >= 3 and out[0] != out[-1]:
        out.append(out[0])
    return out


def shape_to_geojson(shape: Any) -> Optional[dict[str, Any]]:
    parts = getattr(shape, "parts", None)
    pts = list(getattr(shape, "points", []) or [])
    if not pts:
        return None
    if not parts:
        ring = _ring_to_lonlat(pts)
        if len(ring) < 4:
            return None
        return {"type": "Polygon", "coordinates": [ring]}
    rings: list[list[list[float]]] = []
    starts = list(parts) + [len(pts)]
    for i in range(len(starts) - 1):
        ring = _ring_to_lonlat(pts[starts[i] : starts[i + 1]])
        if len(ring) >= 4:
            rings.append(ring)
    if not rings:
        return None
    if len(rings) == 1:
        return {"type": "Polygon", "coordinates": rings}
    return {"type": "MultiPolygon", "coordinates": [[r] for r in rings]}


def geom_bbox(geom: dict[str, Any]) -> Optional[tuple[float, float, float, float]]:
    xs: list[float] = []
    ys: list[float] = []

    def walk(coords: Any) -> None:
        if not coords:
            return
        if isinstance(coords[0], (int, float)):
            xs.append(float(coords[0]))
            ys.append(float(coords[1]))
            return
        for c in coords:
            walk(c)

    walk(geom.get("coordinates"))
    if not xs:
        return None
    return min(xs), min(ys), max(xs), max(ys)


def load_polygons(shp_path: Path) -> list[dict[str, Any]]:
    import shapefile

    r = shapefile.Reader(str(shp_path), encoding="utf-8")
    out: list[dict[str, Any]] = []
    try:
        for rec, shp in zip(r.iterRecords(), r.iterShapes()):
            year, sec_seq, sec_nm, buld_nm = rec[0], rec[1], rec[2], rec[3]
            name = str(sec_nm or "").strip()
            if not name:
                continue
            geom = shape_to_geojson(shp)
            if not geom:
                continue
            bb = geom_bbox(geom)
            out.append(
                {
                    "sec_seq": int(sec_seq),
                    "sec_nm": name,
                    "district_year": int(year) if year else 2024,
                    "buld_nm": str(buld_nm or "").strip(),
                    "geom": geom,
                    "bbox": bb,
                }
            )
    finally:
        r.close()
    return out


def _header_roles(cells: list[Any]) -> Optional[tuple[dict[str, int], list[tuple[int, int, int]]]]:
    roles: dict[str, int] = {}
    quarters: list[tuple[int, int, int]] = []
    for i, raw in enumerate(cells):
        h = str(raw or "").replace("\n", "").strip()
        if h.startswith("지역구분(1)"):
            roles["sido"] = i
        elif h.startswith("지역구분(2)"):
            roles["name"] = i
        elif h.startswith("층구분"):
            roles["floor"] = i
        elif h == "항목":
            roles["item"] = i
        elif h in {"코드", "지역CODE"}:
            roles["code"] = i
        else:
            qh = parse_quarter_header(h)
            if qh:
                quarters.append((i, qh[0], qh[1]))
    if "sido" not in roles or "name" not in roles or "item" not in roles or not quarters:
        return None
    return roles, quarters


def parse_workbook(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    rows_out: list[dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS or sheet_name not in SHEET_ASSET:
                continue
            asset = SHEET_ASSET[sheet_name]
            floor_band = SHEET_FLOOR_BAND.get(sheet_name, "all")
            ws = wb[sheet_name]
            roles = None
            quarters: list[tuple[int, int, int]] = []
            for row in ws.iter_rows(values_only=True):
                cells = list(row)
                if roles is None:
                    parsed = _header_roles(cells)
                    if parsed:
                        roles, quarters = parsed
                    continue
                sido = str(cells[roles["sido"]] or "").strip() if roles["sido"] < len(cells) else ""
                name = str(cells[roles["name"]] or "").strip() if roles["name"] < len(cells) else ""
                item = str(cells[roles["item"]] or "").strip() if roles["item"] < len(cells) else ""
                if is_aggregate_name(name) or not sido:
                    continue
                metric = metric_from_item(item)
                if metric is None:
                    continue
                code = ""
                if "code" in roles and roles["code"] < len(cells) and cells[roles["code"]] is not None:
                    code = str(cells[roles["code"]]).strip()
                floor_label = ""
                if "floor" in roles and roles["floor"] < len(cells):
                    floor_label = str(cells[roles["floor"]] or "").strip()
                for col_i, year, qtr in quarters:
                    if col_i >= len(cells):
                        continue
                    val = parse_number(cells[col_i])
                    if val is None:
                        continue
                    rows_out.append(
                        {
                            "sec_nm": name,
                            "sido": sido,
                            "reb_code": code,
                            "asset_kind": asset,
                            "metric": metric,
                            "floor_band": floor_band,
                            "floor_label": floor_label,
                            "year": year,
                            "quarter": qtr,
                            "value": val,
                            "source_sheet": sheet_name,
                        }
                    )
    finally:
        wb.close()
    return rows_out


def _sido_by_name(facts: list[dict[str, Any]]) -> dict[str, str]:
    out: dict[str, str] = {}
    for r in facts:
        if r["sec_nm"] not in out and r["sido"]:
            out[r["sec_nm"]] = r["sido"]
    return out


def replace_tables(engine, polygons: list[dict[str, Any]], facts: list[dict[str, Any]], source_file: str) -> None:
    sido_map = _sido_by_name(facts)
    poly_rows = []
    for p in polygons:
        bb = p.get("bbox") or (None, None, None, None)
        poly_rows.append(
            (
                p["sec_seq"],
                p["sec_nm"],
                p["district_year"],
                p["buld_nm"],
                sido_map.get(p["sec_nm"], ""),
                Json(p["geom"]),
                bb[0],
                bb[1],
                bb[2],
                bb[3],
            )
        )
    by_grain: dict[tuple, dict[str, Any]] = {}
    for r in facts:
        grain = (
            r["sec_nm"],
            r["asset_kind"],
            r["metric"],
            r["floor_band"],
            r["floor_label"],
            r["year"],
            r["quarter"],
        )
        by_grain[grain] = r
    facts = list(by_grain.values())
    fact_rows = [
        (
            r["sec_nm"],
            r["sido"],
            r["reb_code"],
            r["asset_kind"],
            r["metric"],
            r["floor_band"],
            r["floor_label"],
            r["year"],
            r["quarter"],
            r["value"],
            r["source_sheet"],
        )
        for r in facts
    ]
    years = [r["year"] for r in facts]
    latest_year = max(years) if years else None
    latest_q = max((r["quarter"] for r in facts if r["year"] == latest_year), default=None)

    raw = engine.raw_connection()
    try:
        cur = raw.cursor()
        cur.execute("TRUNCATE rent_sangkwon_quarterly, rent_sangkwon, rent_sangkwon_import_meta")
        execute_values(
            cur,
            """
            INSERT INTO rent_sangkwon (
                sec_seq, sec_nm, district_year, buld_nm, sido,
                geom_geojson, bbox_west, bbox_south, bbox_east, bbox_north
            ) VALUES %s
            """,
            poly_rows,
        )
        execute_values(
            cur,
            """
            INSERT INTO rent_sangkwon_quarterly (
                sec_nm, sido, reb_code, asset_kind, metric, floor_band, floor_label,
                year, quarter, value, source_sheet
            ) VALUES %s
            """,
            fact_rows,
            page_size=2000,
        )
        cur.execute(
            """
            INSERT INTO rent_sangkwon_import_meta (
                id, source_file, latest_year, latest_quarter, n_polygons, n_quarterly, imported_at
            ) VALUES (1, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (id) DO UPDATE SET
                source_file = EXCLUDED.source_file,
                latest_year = EXCLUDED.latest_year,
                latest_quarter = EXCLUDED.latest_quarter,
                n_polygons = EXCLUDED.n_polygons,
                n_quarterly = EXCLUDED.n_quarterly,
                imported_at = NOW()
            """,
            (source_file, latest_year, latest_q, len(poly_rows), len(fact_rows)),
        )
        raw.commit()
    finally:
        raw.close()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=None)
    ap.add_argument("--shp", type=Path, default=DEFAULT_SHP)
    args = ap.parse_args()
    files: list[Path]
    if args.xlsx is not None:
        files = [args.xlsx]
    else:
        found = list(DEFAULT_XLSX_DIR.glob("*.xlsx"))
        if not found:
            log.error("xlsx not found in %s", DEFAULT_XLSX_DIR)
            return 1
        files = [max(found, key=lambda p: p.stat().st_mtime)]
    if not args.shp.exists():
        log.error("shp not found: %s", args.shp)
        return 1

    log.info("polygons %s", args.shp)
    polygons = load_polygons(args.shp)
    log.info("polygons n=%s", len(polygons))
    facts: list[dict[str, Any]] = []
    for xlsx in files:
        log.info("xlsx %s", xlsx)
        part = parse_workbook(xlsx)
        log.info("  quarterly rows %s", len(part))
        facts.extend(part)
    log.info("quarterly rows total %s", len(facts))
    if not polygons or not facts:
        log.error("empty import")
        return 1

    ddl = REPO / "db" / "059_rent_sangkwon.sql"
    engine = get_rent_engine()
    with engine.begin() as conn:
        conn.execute(text(ddl.read_text(encoding="utf-8")))
    source_name = " + ".join(p.name for p in files)
    replace_tables(engine, polygons, facts, source_name)
    log.info("imported %s polygons, %s quarterly", len(polygons), len(facts))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
